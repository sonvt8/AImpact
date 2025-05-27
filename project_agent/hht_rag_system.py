import os
import asyncio
import streamlit as st
import chromadb
import io
import re
import numpy as np
import logging
import unicodedata
import pdfplumber
import hashlib
import pytesseract
import aiohttp
import fasttext
import pandas as pd
import openpyxl
from sqlalchemy import create_engine, text
from langchain_core.callbacks import StreamingStdOutCallbackHandler
from langchain_community.document_loaders import TextLoader, CSVLoader, Docx2txtLoader
from langchain_community.document_loaders.merge import MergedDataLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from underthesea import word_tokenize
from typing import List, Tuple
from datetime import datetime
from functools import lru_cache
from langchain.prompts import PromptTemplate
from langchain_ollama import ChatOllama
from langchain_community.chat_message_histories import SQLChatMessageHistory
from cryptography.fernet import Fernet
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from retry import retry
from dotenv import load_dotenv

try:
    from langchain_openai import ChatOpenAI
except ImportError:
    ChatOpenAI = None

# Thiết lập tùy chọn để tránh cảnh báo từ pandas
pd.set_option('future.no_silent_downcasting', True)

# Tải biến môi trường
load_dotenv()

# Cấu hình hệ thống
DATA_DIR = os.getenv("DATA_DIR", "data")
HISTORY_DIR = os.getenv("HISTORY_DIR", "history")
DOCUMENTS_DIR = os.getenv("DOCUMENTS_DIR", "documents")
CHROMA_DB_PATH = os.path.join(DATA_DIR, "chroma_db")
ENCRYPTION_KEY_PATH = os.path.join(DATA_DIR, "encryption_key.key")
HISTORY_DB_PATH = os.path.join(HISTORY_DIR, "query_history.db")
PASSWORD = os.getenv("APP_PASSWORD", "T0mmy")
SIMILARITY_THRESHOLD_DEFAULT = float(os.getenv("SIMILARITY_THRESHOLD", 0.5))
MODEL_PATH = os.getenv("MODEL_PATH", "./models/cc.vi.300.bin")
CHUNK_SIZE = 1500
CHUNK_OVERLAP = 300
SUPPORTED_FILE_TYPES = [".txt", ".pdf", ".docx", ".xlsx", ".csv"]
CONTEXT_WINDOW_SIZE = 5
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")

# Danh sách từ khóa viễn thông
TELECOM_KEYWORDS = [
    "ACB", "ATS", "MPĐ", "HĐB", "ĐHCX", "UPS", "DC", "UDB", "PDU", "UDB/PDU",
    "LV1", "LV2", "LV3", "LV4", "ESDB2", "ESDB3", "ACDB2_ext", "ACDB3_ext", "MSB4",
    "AMTS1", "AMTS3", "Interlock", "MCCB"
]
KEYWORD_MAP = {keyword: f"KEYWORD_{hashlib.md5(keyword.encode('utf-8')).hexdigest()}" for keyword in TELECOM_KEYWORDS}

# Ánh xạ thuật ngữ chuẩn
TERMINOLOGY_MAPPING = {
    "máy cắt không khí": "ACB",
    "tủ LV1": "tủ điện hạ thế 1",
    "ATS1": "hệ thống chuyển nguồn tự động 1",
    "AMTS1": "hệ thống chuyển nguồn tự động 1",
    "máy phát điện": "MPĐ",
    "điều hòa chính xác": "ĐHCX",
    "tủ phân phối nguồn": "UDB/PDU",
    "sự cố": "lỗi",
    "mất điện": "sự cố lộ điện",
    "mất một lộ điện": "sự cố 1 lộ điện lưới",
    "tự động vận hành": "tự động interlock",
    "xác nhận tình huống": "báo cáo tình huống ƯCTT",
    "nhân viên cơ điện": "NV1 NV2",
    "tòa nhà N6": "N6",
    "tòa nhà N4": "N4",
    "lộ nổi": "lộ nổi",
    "lộ ngầm": "lộ ngầm",
    "máy cắt khối": "MCCB",
    "main power distribution": "MPĐ"
}

# Thiết lập ghi log
logging.basicConfig(level=logging.INFO, filename='rag_system.log', filemode='a',
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Tạo các thư mục cần thiết
for dir_path in [DATA_DIR, HISTORY_DIR, DOCUMENTS_DIR]:
    os.makedirs(dir_path, exist_ok=True)

async def check_ollama():
    # Kiểm tra trạng thái server Ollama
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get("http://localhost:11434", timeout=5) as response:
                return response.status == 200
    except Exception as e:
        logger.error(f"Lỗi kiểm tra server Ollama: {str(e)}")
        return False

async def check_internet():
    # Kiểm tra kết nối internet
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get("https://www.google.com", timeout=5) as response:
                return response.status == 200
    except Exception:
        return False

def display_message(message: str, message_type: str = "error"):
    # Hiển thị thông báo trên giao diện
    if "message_placeholder" not in st.session_state:
        st.session_state.message_placeholder = st.empty()
    if "last_message" not in st.session_state or st.session_state.last_message != message:
        st.session_state.last_message = message
        message_id = f"message_{hashlib.md5(message.encode()).hexdigest()}"
        color = {"error": "#ff4b4b", "warning": "#ffcc00", "info": "#28a745"}.get(message_type, "#ff4b4b")
        html_message = f"""
        <div id="{message_id}" style="padding: 10px; margin-bottom: 10px; border-radius: 5px; color: white; 
        background-color: {color};">
            {message}
        </div>
        <script>
            setTimeout(function() {{
                var elem = document.getElementById("{message_id}");
                if (elem) {{
                    elem.parentNode.removeChild(elem);
                }}
            }}, 5000);
        </script>
        """
        st.session_state.message_placeholder.markdown(html_message, unsafe_allow_html=True)

def check_existing_data(persist_directory: str = CHROMA_DB_PATH) -> bool:
    # Kiểm tra dữ liệu hiện có trong ChromaDB
    try:
        if not os.path.exists(persist_directory):
            return False
        client = chromadb.PersistentClient(path=persist_directory)
        collection = client.get_or_create_collection(name="viettel_docs")
        return collection.count() > 0
    except Exception as e:
        logger.error(f"Lỗi kiểm tra dữ liệu: {str(e)}")
        return False

def preprocess_text(text: str, restore: bool = False, for_embedding: bool = False) -> str:
    # Chuẩn hóa văn bản
    if not text or not isinstance(text, str):
        return ""

    # Chuẩn hóa dạng NFC
    text = unicodedata.normalize('NFC', text)
    text = text.encode('utf-8', 'ignore').decode('utf-8')

    # Xóa ký tự không mong muốn, giữ tiếng Việt
    text = re.sub(r'\.\.\.+', ' ', text)
    for keyword in TELECOM_KEYWORDS:
        text = re.sub(rf'\b{keyword}\s*\d+[aA]\b', lambda m: m.group(0), text, flags=re.IGNORECASE)
    text = re.sub(r'[^\w\s,.?!:;()\-/áàảãạăắằẳẵặâấầẩẫậéèẻẽẹêếềểễệíìỉĩịóòỏõọôốồổỗộơớờởỡợúùủũụưứừửữựýỳỷỹỵÁÀẢÃẠĂẮẰẲẴẶÂẤẦẨẪẬÉÈẺẼẸÊẾỀỂỄỆÍÌỈĨỊÓÒỎÕỌÔỐỒỔỖỘƠỚỜỞỠỢÚÙỦŨỤƯỨỪỬỮỰÝỲỶỸỴ\d]', ' ', text)
    text = re.sub(r'_+', ' ', text)
    text = re.sub(r'\b[-=]{3,}\b', ' ', text)
    text = re.sub(r'\|{2,}', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()

    # Áp dụng ánh xạ thuật ngữ
    for term, standard in TERMINOLOGY_MAPPING.items():
        text = re.sub(rf'\b{term}\b', standard, text, flags=re.IGNORECASE)

    # Xử lý cho embedding hoặc khôi phục từ khóa
    if for_embedding:
        tokens = word_tokenize(text, format="text").split()
        max_tokens = 500
        if len(tokens) > max_tokens:
            tokens = tokens[:max_tokens]
        text = " ".join(tokens)
    elif restore:
        for keyword, placeholder in KEYWORD_MAP.items():
            text = text.replace(placeholder, keyword)
        logger.info(f"Đã khôi phục từ khóa: {text[:100]}...")

    return text

class DocumentProcessor:
    def __init__(self, chunk_size: int = CHUNK_SIZE, chunk_overlap: int = CHUNK_OVERLAP):
        # Khởi tạo bộ chia văn bản
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size, chunk_overlap=chunk_overlap,
            separators=["\n\n", "\n", ".", "!", "?", ",", " ", ""],
        )

    def find_column(self, headers, possible_names):
        # Tìm cột theo danh sách tên có thể
        for name in possible_names:
            for header in headers:
                if name.lower() in str(header).lower():
                    return header
        return None

    def format_row(self, row, headers, sheet_name: str):
        # Định dạng hàng dữ liệu từ Excel
        stt_col = self.find_column(headers, ['Stt', 'Số thứ tự', 'STT'])
        if not stt_col:
            logger.warning(f"Không tìm thấy cột Stt: {headers}")
            return None
        
        stt_value = row.get(stt_col)
        if pd.isna(stt_value):
            logger.info(f"Bỏ qua hàng thiếu Stt: {row.to_dict()}")
            return None
        
        if sheet_name.lower() == "tong hop":
            su_co_col = self.find_column(headers, ['SỰ CỐ'])
            if not su_co_col:
                logger.warning(f"Không tìm thấy cột SỰ CỐ: {headers}")
                return None
            
            su_co_value = row.get(su_co_col)
            if pd.isna(su_co_value):
                logger.info(f"Bỏ qua hàng thiếu SỰ CỐ: {row.to_dict()}")
                return None
            
            formatted = f"Sự cố {su_co_value}: "
            for header in headers:
                if header in ['STT', 'SỰ CỐ']:
                    continue
                value = str(row.get(header, '')) if pd.notna(row.get(header)) else '0'
                formatted += f"{header} {value}, "
            return formatted.rstrip(", ").strip()
        
        situation_col = self.find_column(headers, ['Tình huống', 'Tình huống sự cố', 'Tình huống xử lý', 'Tình huống ƯCTT', 'Tên tình huống', 'Tên lỗi, sự cố'])
        if not situation_col:
            situation_col = self.find_column(headers, ['Mức độ ảnh hưởng', 'Phương án', 'Tình huống sự cố', 'SỰ CỐ'])
            if not situation_col:
                logger.warning(f"Không tìm thấy cột Tình huống: {headers}")
                return None
        
        situation_value = row.get(situation_col)
        if pd.isna(situation_value):
            logger.info(f"Bỏ qua hàng thiếu Tình huống: {row.to_dict()}")
            return None
        
        formatted = f"Tình huống số: {stt_value}\n"
        for header in headers:
            if 'Unnamed' in str(header):
                continue
            value = str(row.get(header, '')) if pd.notna(row.get(header)) else ''
            formatted += f"{header}: {value}\n"
        return formatted.strip()

    def split_long_row(self, row_text, max_chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
        # Chia hàng văn bản dài thành các đoạn nhỏ
        sub_chunks = []
        if len(row_text) <= max_chunk_size:
            sub_chunks.append(row_text)
        else:
            start = 0
            while start < len(row_text):
                end = start + max_chunk_size
                if end < len(row_text):
                    while end > start and row_text[end] not in ['\n', ' ']:
                        end -= 1
                sub_chunk = row_text[start:end]
                sub_chunks.append(sub_chunk)
                start = end - overlap
        return sub_chunks

    def find_header_row(self, filepath, sheet_name):
        # Tìm hàng tiêu đề trong file Excel
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook[sheet_name]
        
        skip_rows = 0
        for row in sheet.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and 'stt' in str(cell.value).lower():
                    return skip_rows
            skip_rows += 1
        return skip_rows

    def process_excel_with_openpyxl(self, filepath, sheet_name):
        # Xử lý file Excel bằng openpyxl
        skip_rows = self.find_header_row(filepath, sheet_name)
        logger.info(f"Sheet {sheet_name}: Bỏ qua {skip_rows} hàng để tìm tiêu đề.")
        
        df = pd.read_excel(filepath, sheet_name=sheet_name, engine='openpyxl', skiprows=skip_rows)
        
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook[sheet_name]
        merged_cells = sheet.merged_cells.ranges
        
        df = df.astype(object)
        
        for merged_range in merged_cells:
            min_row, min_col, max_row, max_col = merged_range.bounds
            value = sheet.cell(min_row, min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    row_idx = row - skip_rows - 1
                    col_idx = col - 1
                    if 0 <= row_idx < len(df) and 0 <= col_idx < len(df.columns):
                        df.iloc[row_idx, col_idx] = value if pd.isna(df.iloc[row_idx, col_idx]) else df.iloc[row_idx, col_idx]
        
        # Chuẩn hóa tên cột
        df.columns = [f"Column_{i}" if str(col).startswith("Unnamed") else col for i, col in enumerate(df.columns)]
        df = df.dropna(how='all').dropna(axis=1, how='all')
        df = df.ffill().infer_objects(copy=False)
        
        # Đảm bảo cột Stt tồn tại
        if 'Stt' not in df.columns and '1' in df.columns:
            df = df.rename(columns={'1': 'Stt'})
        
        return df

    async def load_and_split(self, file, filename: str) -> List[Tuple[List[str], List[int], List[str], str]]:
        # Tải và chia nhỏ tài liệu
        saved_filepath = os.path.join(DOCUMENTS_DIR, filename)
        if not os.path.exists(saved_filepath):
            with open(saved_filepath, "wb") as f:
                f.write(file.read())
        
        loaders = []
        ext = os.path.splitext(filename)[1].lower()
        try:
            if ext == ".txt":
                loaders.append(TextLoader(saved_filepath))
            elif ext == ".pdf":
                return [(await self._process_pdf(saved_filepath), filename)]
            elif ext == ".xlsx":
                try:
                    xl = pd.ExcelFile(saved_filepath, engine='openpyxl')
                    results = []
                    for sheet_name in xl.sheet_names:
                        df = self.process_excel_with_openpyxl(saved_filepath, sheet_name)
                        df = df.dropna(how='all').dropna(axis=1, how='all')
                        df = df.ffill().infer_objects(copy=False)
                        df.columns = df.columns.astype(str).fillna("")
                        if '1' in df.columns:
                            df = df.rename(columns={'1': 'Stt'})
                        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                        if df.empty:
                            logger.info(f"Sheet {sheet_name} rỗng sau xử lý.")
                            continue
                        headers = df.columns.tolist()
                        logger.info(f"Tiêu đề sheet {sheet_name}: {headers}")
                        chunks = []
                        situation_ids = []
                        page_numbers = []
                        for _, row in df.iterrows():
                            logger.info(f"Dữ liệu hàng: {row.to_dict()}")
                            chunk = self.format_row(row, headers, sheet_name)
                            if not chunk:
                                continue
                            logger.info(f"Đoạn thô: {chunk[:100]}...")
                            chunk = preprocess_text(chunk, restore=False)
                            logger.info(f"Đoạn đã xử lý: {chunk[:100]}...")
                            sub_chunks = self.split_long_row(chunk)
                            chunks.extend(sub_chunks)
                            stt_col = self.find_column(headers, ['Stt', 'Số thứ tự', 'STT'])
                            situation_ids.extend([str(row.get(stt_col, ""))] * len(sub_chunks))
                            page_numbers.extend([1] * len(sub_chunks))
                        logger.info(f"Đã xử lý sheet {sheet_name}: {len(chunks)} đoạn")
                        results.append((chunks, page_numbers, situation_ids, sheet_name))
                    return results
                except Exception as e:
                    logger.error(f"Lỗi xử lý file Excel {filename}: {str(e)}")
                    display_message(f"Không thể xử lý file Excel {filename}: {str(e)}", "error")
                    return []
            elif ext == ".csv":
                loaders.append(CSVLoader(saved_filepath))
            else:
                logger.error(f"Loại file không hỗ trợ: {ext}")
                return []
        except Exception as e:
            logger.error(f"Lỗi tải file {filename}: {str(e)}")
            return []

        if loaders:
            merged_loader = MergedDataLoader(loaders=loaders)
            docs = merged_loader.load()
            full_text = " ".join(doc.page_content for doc in docs)
            full_text = preprocess_text(full_text, restore=False)
            splits = self.text_splitter.create_documents([full_text])
            chunks = [chunk.page_content.strip() for chunk in splits if chunk.page_content.strip()]
            page_numbers = [1] * len(chunks)
            situation_ids = [""] * len(chunks)
            return [(chunks, page_numbers, situation_ids, filename)]
        return []

    async def _process_pdf(self, filepath: str) -> Tuple[List[str], List[int], List[str]]:
        # Xử lý file PDF
        full_text = ""
        page_boundaries = []
        current_pos = 0
        try:
            with pdfplumber.open(filepath) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    text = page.extract_text() or ""
                    if not text.strip():
                        logger.info(f"Không trích xuất văn bản từ trang {page_num}, thử OCR...")
                        try:
                            image = page.to_image(resolution=300).original.convert("RGB")
                            text = pytesseract.image_to_string(image, lang='vie').strip()
                        except Exception as e:
                            logger.error(f"Lỗi OCR trang {page_num}: {str(e)}")
                            text = ""
                    tables = page.extract_tables()
                    table_text = "\n".join([",".join(str(cell) if cell is not None else "" for cell in row) for table in tables for row in table]) if tables else ""
                    page_content = (text + "\n" + table_text).strip()
                    if page_content:
                        page_content = preprocess_text(page_content, restore=False)
                        full_text += page_content + "\n"
                        page_boundaries.append((current_pos, page_num))
                        current_pos += len(page_content) + 1
                    else:
                        logger.warning(f"Không có nội dung ở trang {page_num}.")
        except Exception as e:
            logger.error(f"Lỗi xử lý PDF: {str(e)}")
            display_message(f"Lỗi xử lý PDF: {str(e)}", "error")
            return [], [], []
        if not full_text.strip():
            logger.error("Không trích xuất được văn bản từ PDF.")
            return [], [], []
        splits = self.text_splitter.create_documents([full_text])
        chunks = [chunk.page_content.strip() for chunk in splits if chunk.page_content.strip()]
        page_numbers = []
        for chunk in splits:
            chunk_text = chunk.page_content.strip()
            if not chunk_text:
                continue
            chunk_start = full_text.index(chunk_text)
            chunk_page = page_boundaries[0][1] if page_boundaries else 1
            for pos, page_num in page_boundaries:
                if chunk_start >= pos:
                    chunk_page = page_num
                else:
                    break
            page_numbers.append(chunk_page)
        situation_ids = [""] * len(chunks)
        return chunks, page_numbers, situation_ids

class SentenceEmbedding:
    def __init__(self, model_path: str = MODEL_PATH):
        # Khởi tạo mô hình FastText
        if not os.path.exists(model_path):
            logger.info(f"Không tìm thấy mô hình tại {model_path}, đang tải về...")
            os.system("python load_model.py")  # Gọi load_model.py để tải mô hình
            if not os.path.exists(model_path):
                display_message(f"Không thể tải mô hình FastText.", "error")
                st.stop()
        try:
            self.model = fasttext.load_model(model_path)
        except Exception as e:
            logger.error(f"Lỗi tải mô hình FastText từ {model_path}: {str(e)}")
            display_message(f"Lỗi tải mô hình FastText: {str(e)}.", "error")
            st.stop()

    def __call__(self, texts: List[str], is_query: bool = False) -> List[List[float]]:
        # Tạo embedding cho danh sách văn bản
        embeddings = []
        valid_texts = []
        valid_indices = []
        for idx, text in enumerate(texts):
            if not text or not isinstance(text, str):
                logger.warning(f"Bỏ qua văn bản rỗng hoặc không hợp lệ: {text}")
                continue
            try:
                text = preprocess_text(text, for_embedding=True)
                embedding = self.model.get_sentence_vector(text)
                embedding = embedding / np.linalg.norm(embedding)
                embeddings.append(embedding.tolist())
                valid_texts.append(text)
                valid_indices.append(idx)
            except Exception as e:
                logger.error(f"Lỗi tạo embedding cho văn bản: {text[:100]}... Lỗi: {str(e)}")
                continue
        if is_query:
            if not embeddings:
                raise ValueError("Không tạo được embedding cho truy vấn")
            return embeddings[0]
        return embeddings, valid_texts, valid_indices

class ChromaDBManager:
    def __init__(self, persist_directory: str = CHROMA_DB_PATH):
        # Khởi tạo quản lý ChromaDB
        self.key_path = ENCRYPTION_KEY_PATH
        os.makedirs(os.path.dirname(self.key_path), exist_ok=True)
        if os.path.exists(self.key_path):
            with open(self.key_path, "rb") as f:
                self.key = f.read()
        else:
            self.key = Fernet.generate_key()
            with open(self.key_path, "wb") as f:
                f.write(self.key)
        self.cipher = Fernet(self.key)
        os.makedirs(persist_directory, exist_ok=True)
        self.client = chromadb.PersistentClient(path=persist_directory)
        try:
            self.collection = self.client.get_or_create_collection(name="viettel_docs")
            doc_count = self.collection.count()
            logger.info(f"Khởi tạo bộ sưu tập ChromaDB 'viettel_docs' với {doc_count} tài liệu")
        except Exception as e:
            logger.error(f"Lỗi khởi tạo bộ sưu tập ChromaDB: {str(e)}")
            raise

    def verify_collection(self):
        # Xác minh trạng thái bộ sưu tập
        try:
            doc_count = self.collection.count()
            logger.info(f"Xác minh bộ sưu tập: 'viettel_docs' có {doc_count} tài liệu")
            return doc_count > 0
        except Exception as e:
            logger.error(f"Lỗi xác minh bộ sưu tập: {str(e)}")
            return False

    def reset_collection(self):
        # Đặt lại bộ sưu tập
        try:
            self.client.delete_collection("viettel_docs")
            self.collection = self.client.get_or_create_collection(name="viettel_docs")
            logger.info("Đã đặt lại bộ sưu tập 'viettel_docs'")
        except Exception as e:
            logger.error(f"Lỗi đặt lại bộ sưu tập: {str(e)}")
            raise

    def add(self, texts: List[str], embeddings: List[List[float]], filename: str, page_numbers: List[int], sheet_name: str = "", situation_ids: List[str] = None):
        # Thêm tài liệu vào ChromaDB
        if not texts or not embeddings or not page_numbers:
            logger.error("Danh sách rỗng được cung cấp cho ChromaDBManager.add")
            return
        if len(texts) != len(embeddings) or len(texts) != len(page_numbers) or len(texts) != len(situation_ids or []):
            logger.error(f"Độ dài không khớp: texts={len(texts)}, embeddings={len(embeddings)}, page_numbers={len(page_numbers)}, situation_ids={len(situation_ids or [])}")
            return
        self.collection = self.client.get_or_create_collection(name="viettel_docs")
        existing_metas = self.collection.get(include=["metadatas"])["metadatas"]
        existing_docs = {(meta["filename"], meta["sheet_name"]) for meta in existing_metas if isinstance(meta, dict)}
        if (filename, sheet_name) in existing_docs:
            logger.info(f"Bỏ qua tài liệu trùng lặp: {filename}, sheet {sheet_name}")
            return
        existing_ids = set(self.collection.get(include=[])["ids"])
        new_ids = []
        new_texts = []
        new_embeddings = []
        new_metas = []
        for i, (text, emb, sid) in enumerate(zip(texts, embeddings, situation_ids or [""] * len(texts))):
            doc_id = f"doc_{hashlib.md5((filename + sheet_name + sid + str(i) + str(len(existing_ids))).encode()).hexdigest()}"
            logger.info(f"Tạo doc_id: {doc_id} cho sheet {sheet_name}, sid {sid}, index {i}")
            if doc_id in existing_ids:
                logger.warning(f"Phát hiện doc_id trùng lặp: {doc_id}, bỏ qua...")
                continue
            new_ids.append(doc_id)
            new_texts.append(self.cipher.encrypt(text.encode('utf-8')).decode('utf-8'))
            new_embeddings.append(emb)
            meta = {
                "source": f"Tình huống số {i+1}",
                "filename": filename,
                "page_number": page_numbers[i],
                "upload_date": datetime.now().isoformat(),
                "sheet_name": sheet_name,
                "situation_id": sid,
                "keywords": ",".join([kw for kw in TELECOM_KEYWORDS if kw in text])
            }
            new_metas.append(meta)
            logger.info(f"Thêm metadata: {meta}")
            existing_ids.add(doc_id)
        if new_ids:
            self.collection.add(ids=new_ids, documents=new_texts, metadatas=new_metas, embeddings=new_embeddings)
            logger.info(f"Thêm {len(new_ids)} đoạn mới vào ChromaDB")
        else:
            logger.warning("Không có đoạn mới được thêm vào ChromaDB")

    def query(self, query_embedding: List[float], top_k: int, query_text: str = "", is_summary_query: bool = False):
        # Truy vấn tài liệu từ ChromaDB
        if not self.verify_collection():
            logger.info("Bộ sưu tập rỗng hoặc không truy cập được")
            return {"documents": [[]], "metadatas": [[]], "distances": [[]]}

        try:
            res = self.collection.query(
                query_embeddings=[query_embedding],
                n_results=top_k * 2,
                include=["documents", "metadatas", "distances"]
            )
        except Exception as e:
            logger.error(f"Lỗi thực hiện truy vấn ChromaDB: {str(e)}")
            return {"documents": [[]], "metadatas": [[]], "distances": [[]]}

        if not res.get("documents") or not isinstance(res["documents"], list) or len(res["documents"]) == 0:
            logger.info("Không tìm thấy tài liệu cho truy vấn")
            return {"documents": [[]], "metadatas": [[]], "distances": [[]]}

        if not res["documents"][0]:
            logger.info("Danh sách tài liệu rỗng")
            return {"documents": [[]], "metadatas": [[]], "distances": [[]]}

        docs = res["documents"][0]
        metas = res["metadatas"][0]
        distances = res["distances"][0]

        if not (len(docs) == len(metas) == len(distances)):
            logger.error(f"Độ dài không khớp: docs={len(docs)}, metas={len(metas)}, distances={len(distances)}")
            return {"documents": [[]], "metadatas": [[]], "distances": [[]]}

        valid_docs = []
        valid_metas = []
        valid_distances = []
        for doc, meta, dist in zip(docs, metas, distances):
            if not isinstance(meta, dict):
                logger.error(f"Định dạng metadata không hợp lệ: meta={meta}, type={type(meta)}")
                continue
            valid_docs.append(doc)
            valid_metas.append(meta)
            valid_distances.append(dist)

        if not valid_docs:
            logger.info("Không tìm thấy tài liệu hợp lệ")
            return {"documents": [[]], "metadatas": [[]], "distances": [[]]}

        docs = valid_docs
        metas = valid_metas
        distances = valid_distances

        distances = np.array(distances)
        similarities = 1 - distances

        # Lọc theo tòa nhà và hệ thống
        query_lower = query_text.lower()
        building = "N4" if "n4" in query_lower else "N6" if "n6" in query_lower else None
        systems = [kw for kw in TELECOM_KEYWORDS if kw.lower() in query_lower]
        query_keywords = set(word.lower() for word in query_text.split() if word.lower() in [kw.lower() for kw in TELECOM_KEYWORDS])

        filtered_docs = []
        filtered_metas = []
        filtered_similarities = []

        for i, (doc, meta, sim) in enumerate(zip(docs, metas, similarities)):
            sheet_name = meta.get("sheet_name", "").lower()
            if is_summary_query and sheet_name != "tong hop":
                continue
            if not is_summary_query and sheet_name not in ["1.ds uctt", "form xlsc"]:
                continue

            doc_content = self.cipher.decrypt(doc.encode('utf-8')).decode('utf-8').lower()
            building_match = True
            if building and building.lower() not in doc_content and building.lower() not in meta.get("situation_id", "").lower():
                logger.info(f"Lọc bỏ tài liệu do không khớp tòa nhà: {doc_content[:100]}...")
                building_match = False

            doc_has_system = False
            if systems:
                for system in systems:
                    core_system = re.sub(r'\d+a', '', system.lower())
                    full_system = system.lower()
                    if (core_system and core_system in doc_content) or (full_system in doc_content):
                        doc_has_system = True
                        logger.info(f"Khớp hệ thống: {system}, tài liệu: {doc_content[:100]}...")
                        break
                if not doc_has_system:
                    logger.info(f"Lọc bỏ tài liệu do không khớp hệ thống: {doc_content[:100]}...")
            else:
                doc_has_system = True

            if building_match and doc_has_system:
                filtered_docs.append(doc)
                filtered_metas.append(meta)
                doc_keywords = set(word.lower() for word in doc_content.split())
                meta_keywords = set(kw.lower() for kw in meta.get("keywords", "").split(",") if kw)
                matched_keywords = query_keywords & (doc_keywords | meta_keywords)
                boost = 0.1 * len(matched_keywords)
                filtered_similarities.append(sim + boost)

        if not filtered_docs:
            logger.info("Không có tài liệu nào vượt qua bộ lọc")
            return {"documents": [[]], "metadatas": [[]], "distances": [[]]}

        sorted_indices = np.argsort(-np.array(filtered_similarities))[:top_k]
        restored_docs = [
            preprocess_text(
                self.cipher.decrypt(filtered_docs[i].encode('utf-8')).decode('utf-8'),
                restore=True
            ) for i in sorted_indices
        ]
        res = {
            "documents": [restored_docs],
            "metadatas": [[filtered_metas[i] for i in sorted_indices]],
            "distances": [[filtered_similarities[i] for i in sorted_indices]]
        }

        logger.info(f"Độ tương đồng cosine: {res['distances']}")
        return res

    def list_documents(self) -> List[str]:
        # Liệt kê tài liệu đã tải
        try:
            results = self.collection.get(include=["metadatas"])
            return list(set(meta["filename"] for meta in results["metadatas"]))
        except Exception as e:
            display_message(f"Lỗi liệt kê tài liệu: {str(e)}", "error")
            return []

class AnswerGenerator:
    def __init__(self, role: str = "Expert"):
        # Khởi tạo trình tạo câu trả lời
        self.role = role
        self.model = None
        self.model_type = None

    async def initialize(self, model_type: str):
        # Khởi tạo mô hình LLM
        self.model_type = model_type
        if model_type == "ollama":
            if not await check_ollama():
                display_message("Server Ollama không chạy. Chạy 'ollama run llama3.2'.", "error")
                st.stop()
            self.model = ChatOllama(
                base_url="http://localhost:11434",
                model="llama3.2",
                temperature=0.1,
                max_tokens=2000,
                streaming=True,
                callbacks=[StreamingStdOutCallbackHandler()]
            )
        elif model_type == "openai":
            if not ChatOpenAI:
                display_message("Mô-đun langchain-openai chưa cài đặt.", "error")
                st.stop()
            if not OPENAI_API_KEY:
                display_message("Chưa thiết lập khóa API OpenAI.", "error")
                st.stop()
            if not await check_internet():
                display_message("Không có kết nối internet. Chuyển sang Ollama.", "warning")
                if not await check_ollama():
                    display_message("Server Ollama không chạy.", "error")
                    st.stop()
                self.model_type = "ollama"
                self.model = ChatOllama(
                    base_url="http://localhost:11434",
                    model="llama3.2",
                    temperature=0.1,
                    max_tokens=2000,
                    streaming=True,
                    callbacks=[StreamingStdOutCallbackHandler()]
                )
            else:
                self.model = ChatOpenAI(
                    model="gpt-4o-mini",
                    api_key=OPENAI_API_KEY,
                    temperature=0.1,
                    max_tokens=2000,
                    streaming=True,
                    callbacks=[StreamingStdOutCallbackHandler()]
                )
        else:
            raise ValueError(f"Loại mô hình không hỗ trợ: {model_type}")

    @retry(tries=3, delay=1, backoff=2)
    async def generate_answer_stream(self, question: str, context: str, citations: str, conversation_history: str, is_summary_query: bool = False):
        # Tạo câu trả lời dưới dạng stream
        if is_summary_query:
            prompt = PromptTemplate(
                input_variables=["conversation_history", "question", "context", "citations", "role"],
                template="""Bạn là chuyên gia viễn thông và vận hành Data Center, trả lời ở mức độ {role}. Dựa trên câu hỏi và ngữ cảnh, cung cấp câu trả lời chính xác về thống kê sự cố trong năm qua.

                    Câu hỏi: {question}

                    Ngữ cảnh: {context}

                    Hướng dẫn:
                    - Chỉ dùng thông tin từ ngữ cảnh.
                    - Trả lời về số lượng sự cố (ƯCTT, XLSC, VHKT, TỔNG) từ sheet "Tong hop".
                    - Nếu không có thông tin, trả lời: "Không tìm thấy thông tin thống kê phù hợp."
                    - Trả lời rõ ràng, ngắn gọn, đúng dữ liệu.
                    - Không suy diễn ngoài tài liệu.

                    Trả lời:"""
            )
        else:
            prompt = PromptTemplate(
                input_variables=["conversation_history", "question", "context", "citations", "role"],
                template="""Bạn là chuyên gia viễn thông và vận hành Data Center, trả lời ở mức độ {role}. Dựa trên câu hỏi và ngữ cảnh, cung cấp câu trả lời chính xác theo định dạng:

                    - **Tình huống**: [Mô tả sự cố, bao gồm tòa nhà (N4/N6) và hệ thống (MPĐ, AC, UPS)]
                    - **Dấu hiệu nhận biết**: [Dấu hiệu cụ thể]
                    - **Giải pháp thực hiện**: [Các bước cụ thể]
                    - **Nguyên nhân**: [Nguyên nhân nếu có]
                    - **Mức độ sự cố**: [Lớn/Trung bình/Nhỏ nếu có]
                    - **Nguồn vật tư**: [Nơi lấy vật tư nếu có]
                    - **Ghi chú**: [Thông tin bổ sung]

                    Câu hỏi: {question}

                    Ngữ cảnh: {context}

                    Hướng dẫn:
                    - Dùng thông tin từ ngữ cảnh.
                    - Liên quan trực tiếp đến tòa nhà và hệ thống được hỏi.
                    - Nếu không có thông tin, trả lời: "Không tìm thấy thông tin phù hợp."
                    - Không suy diễn ngoài tài liệu.
                    - Chọn tình huống phù hợp nhất.

                    Trả lời:"""
            )

        try:
            if "Không có thông tin từ tài liệu" in context:
                yield "Không tìm thấy thông tin phù hợp."
                return
            formatted_prompt = prompt.format(
                conversation_history=conversation_history,
                question=question,
                context=context,
                citations=citations,
                role=self.role
            )
            async for chunk in self.model.astream(formatted_prompt):
                yield chunk.content
        except Exception as e:
            display_message(f"Lỗi tạo câu trả lời: {str(e)}", "error")
            yield ""

    @retry(tries=3, delay=1, backoff=2)
    def generate_answer(self, question: str, context: str, citations: str, conversation_history: str, is_summary_query: bool = False) -> str:
        # Tạo câu trả lời đầy đủ
        if is_summary_query:
            prompt = PromptTemplate(
                input_variables=["conversation_history", "question", "context", "citations", "role"],
                template="""Bạn là chuyên gia viễn thông và vận hành Data Center, trả lời ở mức độ {role}. Dựa trên câu hỏi và ngữ cảnh, cung cấp câu trả lời chính xác về thống kê sự cố trong năm qua.

                    Câu hỏi: {question}

                    Ngữ cảnh: {context}

                    Hướng dẫn:
                    - Chỉ dùng thông tin từ ngữ cảnh.
                    - Trả lời về số lượng sự cố (ƯCTT, XLSC, VHKT, TỔNG) từ sheet "Tong hop".
                    - Nếu không có thông tin, trả lời: "Không tìm thấy thông tin thống kê phù hợp."
                    - Trả lời rõ ràng, ngắn gọn, đúng dữ liệu.
                    - Không suy diễn ngoài tài liệu.

                    Trả lời:"""
            )
        else:
            prompt = PromptTemplate(
                input_variables=["conversation_history", "question", "context", "citations", "role"],
                template="""Bạn là chuyên gia viễn thông và vận hành Data Center, trả lời ở mức độ {role}. Dựa trên câu hỏi và ngữ cảnh, cung cấp câu trả lời chính xác theo định dạng:

                    - **Tình huống**: [Mô tả sự cố, bao gồm tòa nhà (N4/N6) và hệ thống (MPĐ, AC, UPS)]
                    - **Dấu hiệu nhận biết**: [Dấu hiệu cụ thể]
                    - **Giải pháp thực hiện**: [Các bước cụ thể]
                    - **Nguyên nhân**: [Nguyên nhân nếu có]
                    - **Mức độ sự cố**: [Lớn/Trung bình/Nhỏ nếu có]
                    - **Nguồn vật tư**: [Nơi lấy vật tư nếu có]
                    - **Ghi chú**: [Thông tin bổ sung]

                    Câu hỏi: {question}

                    Ngữ cảnh: {context}

                    Hướng dẫn:
                    - Dùng thông tin từ ngữ cảnh.
                    - Liên quan trực tiếp đến tòa nhà và hệ thống được hỏi.
                    - Nếu không có thông tin, trả lời: "Không tìm thấy thông tin phù hợp."
                    - Không suy diễn ngoài tài liệu.
                    - Chọn tình huống phù hợp nhất.

                    Trả lời:"""
            )
        
        try:
            if "Không có thông tin từ tài liệu" in context:
                return "Không tìm thấy thông tin phù hợp."
            formatted_prompt = prompt.format(
                conversation_history=conversation_history,
                question=question,
                context=context,
                citations=citations,
                role=self.role
            )
            response = self.model.invoke(formatted_prompt)
            return response.content.strip()
        except Exception as e:
            display_message(f"Lỗi tạo câu trả lời: {str(e)}", "error")
            return ""

def get_session_history(conversation_id: str):
    # Lấy lịch sử cuộc trò chuyện
    try:
        engine = create_engine(f"sqlite:///{HISTORY_DB_PATH}")
        with engine.connect() as connection:
            connection.execute(text("""
                CREATE TABLE IF NOT EXISTS chat_messages (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    session_id TEXT NOT NULL,
                    type TEXT NOT NULL,
                    content TEXT NOT NULL,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
            connection.commit()
        return SQLChatMessageHistory(session_id=conversation_id, connection=f"sqlite:///{HISTORY_DB_PATH}")
    except Exception as e:
        logger.error(f"Lỗi khởi tạo lịch sử: {str(e)}")
        return SQLChatMessageHistory(session_id=conversation_id, connection=f"sqlite:///{HISTORY_DB_PATH}")

def load_conversations():
    # Tải danh sách cuộc trò chuyện
    try:
        engine = create_engine(f"sqlite:///{HISTORY_DB_PATH}")
        with engine.connect() as connection:
            connection.execute(text("""
                CREATE TABLE IF NOT EXISTS chat_messages (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    session_id TEXT NOT NULL,
                    type TEXT NOT NULL,
                    content TEXT NOT NULL,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
            connection.commit()
            result = connection.execute(text("SELECT DISTINCT session_id FROM chat_messages ORDER BY session_id"))
            return [row[0] for row in result]
    except Exception as e:
        logger.error(f"Lỗi tải cuộc trò chuyện: {str(e)}")
        return []

def check_table_exists():
    # Kiểm tra bảng chat_messages tồn tại
    try:
        engine = create_engine(f"sqlite:///{HISTORY_DB_PATH}")
        with engine.connect() as connection:
            result = connection.execute(
                text("SELECT name FROM sqlite_master WHERE type='table' AND name='chat_messages'")
            )
            return result.fetchone() is not None
    except Exception as e:
        logger.error(f"Lỗi kiểm tra bảng: {str(e)}")
        return False

def wrap_text(text: str, width: float, canvas_obj, font_name: str = "Helvetica", font_size: int = 12) -> List[str]:
    # Chia văn bản thành các dòng để hiển thị trong PDF
    lines = []
    current_line = ""
    canvas_obj.setFont(font_name, font_size)
    for word in text.split():
        test_line = current_line + word + " "
        if canvas_obj.stringWidth(test_line, font_name, font_size) <= width:
            current_line = test_line
        else:
            lines.append(current_line.strip())
            current_line = word + " "
    if current_line:
        lines.append(current_line.strip())
    return lines

def export_history_to_pdf(history, font_name: str = "Times-Roman", font_size: int = 12) -> io.BytesIO:
    # Xuất lịch sử truy vấn ra PDF
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Sử dụng font hỗ trợ tiếng Việt
    font_available = os.path.exists("fonts/DejaVuSans.ttf")
    if font_available:
        try:
            pdfmetrics.registerFont(TTFont("DejaVuSans", "fonts/DejaVuSans.ttf"))
            font_name = "DejaVuSans"
            logger.info("Sử dụng font DejaVuSans cho PDF.")
        except Exception as e:
            display_message(f"Lỗi tải font DejaVuSans: {str(e)}. Dùng font mặc định.", "warning")
            font_name = "Times-Roman"
    else:
        font_name = "Times-Roman"
        logger.info("Không tìm thấy DejaVuSans, dùng font Times-Roman.")

    y_position = height - inch
    c.setFont(font_name, 16)
    c.drawString(inch, y_position, "Lịch sử truy vấn".encode('utf-8').decode('utf-8'))
    y_position -= 0.5 * inch
    c.setFont(font_name, font_size)
    
    for i in range(0, len(history), 2):
        if i + 1 < len(history):
            user_msg = history[i].content.encode('utf-8').decode('utf-8')
            ai_msg = history[i + 1].content.encode('utf-8').decode('utf-8')
            c.setFont(font_name, font_size)
            c.drawString(inch, y_position, "Câu hỏi:")
            y_position -= 0.3 * inch
            for line in wrap_text(user_msg, width - 2 * inch, c, font_name, font_size):
                if y_position < inch:
                    c.showPage()
                    c.setFont(font_name, font_size)
                    y_position = height - inch
                c.drawString(inch, y_position, line)
                y_position -= 0.3 * inch
            y_position -= 0.3 * inch
            c.setFont(font_name, font_size)
            c.drawString(inch, y_position, "Câu trả lời:")
            y_position -= 0.3 * inch
            for line in wrap_text(ai_msg, width - 2 * inch, c, font_name, font_size):
                if y_position < inch:
                    c.showPage()
                    c.setFont(font_name, font_size)
                    y_position = height - inch
                c.drawString(inch, y_position, line)
                y_position -= 0.3 * inch
            y_position -= 0.5 * inch
            if y_position < inch:
                c.showPage()
                y_position = height - inch
    c.save()
    buffer.seek(0)
    return buffer

def setup_session_state():
    # Thiết lập trạng thái phiên Streamlit
    if "message_placeholder" not in st.session_state:
        st.session_state.message_placeholder = st.empty()
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "processor" not in st.session_state:
        st.session_state.processor = DocumentProcessor()
    if "embedder" not in st.session_state:
        st.session_state.embedder = SentenceEmbedding()
    if "chroma" not in st.session_state:
        st.session_state.chroma = ChromaDBManager(persist_directory=CHROMA_DB_PATH)
    if "conversations" not in st.session_state:
        st.session_state.conversations = {}
    if "conversation_order" not in st.session_state:
        st.session_state.conversation_order = load_conversations()
    if "current_conversation_id" not in st.session_state:
        conversation_id = f"conversation_{datetime.now().isoformat()}_{hashlib.md5(str(datetime.now()).encode()).hexdigest()}"
        st.session_state.conversations[conversation_id] = get_session_history(conversation_id)
        st.session_state.conversation_order.append(conversation_id)
        st.session_state.current_conversation_id = conversation_id
        logger.info(f"Khởi tạo ID cuộc trò chuyện: {conversation_id}")
    if "has_db_notified" not in st.session_state:
        st.session_state.has_db_notified = False
    if "confirm_delete" not in st.session_state:
        st.session_state.confirm_delete = False
    if "last_answer" not in st.session_state:
        st.session_state.last_answer = None
    if "last_citations" not in st.session_state:
        st.session_state.last_citations = None
    if "query_input_value" not in st.session_state:
        st.session_state.query_input_value = ""

def handle_authentication():
    # Xác thực người dùng
    if not st.session_state.authenticated:
        # Khởi tạo password trong session_state nếu chưa có
        if "password" not in st.session_state:
            st.session_state.password = ""

        # Sử dụng st.form để tránh rerender không cần thiết
        with st.form(key="auth_form"):
            st.session_state.password = st.text_input("Nhập mật khẩu:", type="password", value=st.session_state.password)
            submit_button = st.form_submit_button("Xác thực")

        if submit_button:
            if st.session_state.password == PASSWORD:
                st.session_state.authenticated = True
                logger.info("Xác thực thành công.")
            else:
                display_message("Mật khẩu không đúng.", "error")
        return False
    return True

async def process_files(uploaded_files):
    # Xử lý các file tải lên
    try:
        for uploaded_file in uploaded_files:
            results = await st.session_state.processor.load_and_split(uploaded_file, uploaded_file.name)
            for chunks, page_numbers, situation_ids, sheet_name in results:
                if not chunks:
                    display_message(f"Không trích xuất được văn bản từ {uploaded_file.name}, sheet {sheet_name}.", "error")
                    continue
                total_length = sum(len(chunk) for chunk in chunks)
                if total_length > 1000000:
                    display_message(f"Nội dung {uploaded_file.name} (sheet {sheet_name}) quá lớn.", "warning")
                embeddings, valid_chunks, valid_indices = st.session_state.embedder(chunks)
                if not embeddings or not valid_chunks:
                    display_message(f"Không tạo được embeddings cho {uploaded_file.name} (sheet {sheet_name}).", "error")
                    continue
                valid_page_numbers = [page_numbers[i] for i in valid_indices]
                valid_situation_ids = [situation_ids[i] for i in valid_indices]
                st.session_state.chroma.add(
                    texts=valid_chunks,
                    embeddings=embeddings,
                    filename=uploaded_file.name,
                    page_numbers=valid_page_numbers,
                    sheet_name=sheet_name,
                    situation_ids=valid_situation_ids
                )
        display_message("Xử lý và lưu tài liệu thành công!", "info")
        st.session_state.has_db_notified = False
    except Exception as e:
        logger.error(f"Lỗi xử lý tài liệu: {str(e)}")
        display_message(f"Lỗi xử lý tài liệu: {str(e)}", "error")

def display_documents():
    # Hiển thị danh sách tài liệu
    if "chroma" in st.session_state:
        documents = st.session_state.chroma.list_documents()
        if documents:
            for doc in documents:
                st.sidebar.write(f"- {doc}")
        else:
            st.sidebar.write("Chưa có tài liệu.")
    else:
        st.sidebar.write("Chưa khởi tạo ChromaDB.")

def delete_all_documents():
    # Xóa tất cả tài liệu
    try:
        if "chroma" in st.session_state:
            st.session_state.chroma.reset_collection()
            st.session_state.chroma = ChromaDBManager(persist_directory=CHROMA_DB_PATH)
            st.session_state.embedder = SentenceEmbedding()
            st.session_state.processor = DocumentProcessor()
            for file in os.listdir(DOCUMENTS_DIR):
                file_path = os.path.join(DOCUMENTS_DIR, file)
                if os.path.isfile(file_path):
                    os.remove(file_path)
            display_message("Đã xóa tất cả tài liệu!", "info")
            st.session_state.has_db_notified = False
        else:
            display_message("Chưa khởi tạo ChromaDB.", "error")
    except Exception as e:
        display_message(f"Lỗi xóa tài liệu: {str(e)}", "error")

def create_new_conversation(confirmed=False):
    # Tạo cuộc trò chuyện mới
    if len(st.session_state.conversation_order) >= 5 and not confirmed:
        oldest_conversation = st.session_state.conversation_order[0]
        st.session_state.confirm_delete = True
        display_message(
            f"Đã đạt giới hạn 5 cuộc trò chuyện. Xóa cuộc trò chuyện cũ nhất ({oldest_conversation.split('_')[1]})?",
            "warning"
        )
        return False

    if confirmed and len(st.session_state.conversation_order) >= 5:
        oldest_conversation = st.session_state.conversation_order.pop(0)
        if check_table_exists():
            try:
                engine = create_engine(f"sqlite:///{HISTORY_DB_PATH}")
                with engine.connect() as connection:
                    connection.execute(
                        text("DELETE FROM chat_messages WHERE session_id = :conversation_id"),
                        {"conversation_id": oldest_conversation}
                    )
                    connection.commit()
                del st.session_state.conversations[oldest_conversation]
                display_message(f"Đã xóa cuộc trò chuyện cũ nhất.", "info")
            except Exception as e:
                display_message(f"Lỗi xóa cuộc trò chuyện: {str(e)}", "error")
                return False
        else:
            del st.session_state.conversations[oldest_conversation]
            display_message(f"Không có dữ liệu lịch sử, đã xóa cuộc trò chuyện cũ.", "info")

    conversation_id = f"conversation_{datetime.now().isoformat()}_{hashlib.md5(str(datetime.now()).encode()).hexdigest()}"
    st.session_state.conversations[conversation_id] = get_session_history(conversation_id)
    st.session_state.conversation_order.append(conversation_id)
    st.session_state.current_conversation_id = conversation_id
    st.session_state.confirm_delete = False
    st.session_state.last_answer = None
    st.session_state.last_citations = None
    display_message("Đã tạo cuộc trò chuyện mới!", "info")
    logger.info(f"Tạo ID cuộc trò chuyện mới: {conversation_id}")
    return True

def display_conversation_list():
    # Hiển thị danh sách cuộc trò chuyện
    st.sidebar.subheader("Danh sách cuộc trò chuyện")
    if st.session_state.conversation_order:
        for conversation_id in reversed(st.session_state.conversation_order):
            conversation_time = conversation_id.split("_")[1]
            if st.sidebar.button(f"Cuộc trò chuyện - {conversation_time}", key=conversation_id):
                st.session_state.current_conversation_id = conversation_id
                st.session_state.confirm_delete = False
                st.session_state.last_answer = None
                st.session_state.last_citations = None
                display_message(f"Chuyển sang cuộc trò chuyện {conversation_time}", "info")
                logger.info(f"Chuyển sang ID cuộc trò chuyện: {conversation_id}")
    else:
        st.sidebar.write("Chưa có cuộc trò chuyện.")

def get_conversation_context(conversation_id: str) -> str:
    # Lấy ngữ cảnh cuộc trò chuyện
    if not conversation_id or conversation_id not in st.session_state.conversations:
        return "Không có lịch sử trò chuyện."
    
    messages = st.session_state.conversations[conversation_id].messages
    if not messages:
        return "Không có lịch sử trò chuyện."
    
    recent_messages = messages[-CONTEXT_WINDOW_SIZE:] if len(messages) > CONTEXT_WINDOW_SIZE else messages
    context = ""
    for i in range(0, len(recent_messages), 2):
        if i + 1 < len(recent_messages):
            user_msg = recent_messages[i].content
            ai_msg = recent_messages[i + 1].content
            context += f"Người dùng: {user_msg}\nTrợ lý: {ai_msg}\n"
        else:
            user_msg = recent_messages[i].content
            context += f"Người dùng: {user_msg}\n"
    return context.strip()

async def handle_query(query: str, llm_type: str, role: str, similarity_threshold: float):
    # Xử lý truy vấn người dùng
    if "chroma" not in st.session_state or st.session_state.chroma is None or "embedder" not in st.session_state:
        display_message("Chưa có dữ liệu. Vui lòng tải tài liệu.", "warning")
        return

    try:
        current_conversation_id = st.session_state.current_conversation_id
        logger.info(f"Xử lý truy vấn với ID: {current_conversation_id}")

        conversation_history = get_conversation_context(current_conversation_id)
        logger.info(f"Ngữ cảnh cuộc trò chuyện: {conversation_history}")

        is_summary_query = any(keyword in query.lower() for keyword in ["bao nhiêu sự cố", "tổng số sự cố", "thống kê sự cố", "trong năm vừa qua"])
        expanded_query = query
        for term, standard in TERMINOLOGY_MAPPING.items():
            if term.lower() in query.lower():
                expanded_query += f" {standard}"
        logger.info(f"Truy vấn mở rộng: {expanded_query}")

        answer_generator = AnswerGenerator(role=role)
        await answer_generator.initialize(model_type=llm_type)
        
        context = "Không có thông tin từ tài liệu."
        citations = "Không có trích dẫn."
        citations_list = []
        
        logger.info("Truy vấn cơ sở dữ liệu.")
        @lru_cache(maxsize=100)
        def cached_query(query: str, top_k: int) -> Tuple[List[str], List[dict], List[float]]:
            query = preprocess_text(query, restore=False)
            emb = st.session_state.embedder([query], is_query=True)
            res = st.session_state.chroma.query(emb, top_k, query_text=query, is_summary_query=is_summary_query)
            logger.info(f"Kết quả ChromaDB: documents={len(res['documents']) if res.get('documents') else 0}")
            docs = res.get("documents", [[]])[0]
            metas = res.get("metadatas", [[]])[0]
            dists = res.get("distances", [[]])[0]
            if not isinstance(docs, list) or not isinstance(metas, list) or not isinstance(dists, list):
                logger.error(f"Định dạng dữ liệu không hợp lệ")
                return [], [], []
            if len(docs) != len(metas) or len(docs) != len(dists):
                logger.error(f"Độ dài không khớp")
                return [], [], []
            return docs, metas, dists

        docs, metas, dists = cached_query(expanded_query, top_k=10)
        logger.info(f"Lấy được {len(docs)} tài liệu")
        
        if not docs or not isinstance(docs, list):
            context = "Không có thông tin phù hợp."
            citations = f"Không tìm thấy tài liệu liên quan đến '{query}'."
        else:
            relevant_docs = []
            relevant_metas = []
            relevant_dists = []
            
            for doc, meta, dist in zip(docs, metas, dists):
                if not isinstance(doc, str) or not isinstance(meta, dict) or not isinstance(dist, (int, float)):
                    logger.warning(f"Dữ liệu tài liệu không hợp lệ")
                    continue
                if dist >= similarity_threshold:
                    relevant_docs.append(doc)
                    relevant_metas.append(meta)
                    relevant_dists.append(dist)

            if relevant_docs:
                sorted_pairs = sorted(zip(relevant_docs, relevant_metas, relevant_dists), key=lambda x: x[2], reverse=True)[:5]
                context_parts = []
                for doc, meta, dist in sorted_pairs:
                    if not isinstance(meta, dict):
                        logger.error(f"Định dạng metadata không hợp lệ")
                        continue
                    context_parts.append(
                        f"Tài liệu (Sheet: {meta.get('sheet_name', 'Unknown')}, ID: {meta.get('situation_id', 'Unknown')}, Tương đồng: {dist:.4f}):\n{doc}\n"
                    )
                context = "\n".join(context_parts) if context_parts else "Không có thông tin phù hợp."
                if context_parts:
                    for doc, meta, dist in sorted_pairs:
                        if not isinstance(meta, dict):
                            logger.error(f"Định dạng metadata không hợp lệ")
                            continue
                        citations_list.append({
                            "source": meta.get("source", "Unknown"),
                            "filename": meta.get("filename", "Unknown"),
                            "sheet_name": meta.get("sheet_name", "Không xác định"),
                            "situation_id": meta.get("situation_id", "Không xác định"),
                            "page_number": meta.get("page_number", "Không xác định"),
                            "score": dist,
                            "content": preprocess_text(doc, restore=True)
                        })
                    citations = "Có trích dẫn liên quan."
                else:
                    context = "Không có thông tin phù hợp."
                    citations = f"Không tìm thấy tài liệu đạt ngưỡng tương đồng."
            else:
                context = "Không có thông tin phù hợp."
                citations = f"Không tìm thấy tài liệu đạt ngưỡng tương đồng."

        st.markdown("### Câu trả lời:")
        response_container = st.empty()
        response_text = ""
        async for chunk in answer_generator.generate_answer_stream(query, context, citations, conversation_history, is_summary_query):
            response_text += chunk
            response_container.markdown(response_text)

        if st.session_state.current_conversation_id != current_conversation_id:
            logger.warning(f"ID cuộc trò chuyện thay đổi")
            current_conversation_id = st.session_state.current_conversation_id

        current_history = st.session_state.conversations[current_conversation_id]
        current_history.add_user_message(query)
        current_history.add_ai_message(response_text)
        engine = create_engine(f"sqlite:///{HISTORY_DB_PATH}")
        with engine.connect() as connection:
            connection.commit()
        st.session_state.last_answer = response_text
        st.session_state.last_citations = citations_list
        st.session_state.query_input_value = ""
        logger.info(f"Xử lý truy vấn thành công, ID: {current_conversation_id}")

    except Exception as e:
        error_message = f"Lỗi truy vấn: {str(e)}."
        display_message(error_message, "error")
        logger.error(f"Lỗi truy vấn: {str(e)}")

def display_query_history():
    # Hiển thị lịch sử truy vấn
    if not st.session_state.current_conversation_id:
        st.write("Chưa chọn cuộc trò chuyện.")
        return
    messages = st.session_state.conversations[st.session_state.current_conversation_id].messages
    for i in range(0, len(messages), 2):
        if i + 1 < len(messages):
            user_msg = messages[i]
            ai_msg = messages[i + 1]
            user_content = user_msg.content.replace("\n", "<br>")
            ai_content = ai_msg.content.replace("\n", "<br>")
            st.markdown(
                f"""
                <div style="background-color: #ffffff; padding: 10px; border-radius: 5px; margin-bottom: 10px;">
                    👤 <strong>Người dùng:</strong> {user_content}
                </div>
                <div style="background-color: #d3d3d3; padding: 10px; border-radius: 5px; margin-bottom: 10px;">
                    🤖 <strong>Trợ lý:</strong> {ai_content}
                </div>
                {"---" if i + 2 < len(messages) else ""}
                """,
                unsafe_allow_html=True
            )
        else:
            user_msg = messages[i]
            user_content = user_msg.content.replace("\n", "<br>")
            st.markdown(
                """
                <div style="background-color: #ffffff; padding: 10px; border-radius: 5px; margin-bottom: 10px;">
                    👤 <strong>Người dùng:</strong> {user_content}
                </div>
                """.format(user_content=user_content),
                unsafe_allow_html=True
            )
    if messages and st.button("Xuất lịch sử truy vấn"):
        pdf_buffer = export_history_to_pdf(messages)
        st.download_button(
            "Tải lịch sử truy vấn (PDF)",
            pdf_buffer,
            file_name=f"query_history_{st.session_state.current_conversation_id}.pdf",
            mime="application/pdf"
        )

def display_summary_data():
    # Hiển thị thống kê sự cố
    if "chroma" not in st.session_state:
        st.write("Chưa khởi tạo ChromaDB.")
        return
    
    try:
        results = st.session_state.chroma.collection.get(include=["documents", "metadatas"])
        summary_docs = []
        for doc, meta in zip(results["documents"], results["metadatas"]):
            if meta.get("sheet_name", "").lower() == "tong hop":
                decrypted_doc = st.session_state.chroma.cipher.decrypt(doc.encode('utf-8')).decode('utf-8')
                restored_doc = preprocess_text(decrypted_doc, restore=True)
                logger.info(f"Tài liệu hiển thị: {restored_doc[:100]}...")
                summary_docs.append(restored_doc)
        
        if summary_docs:
            st.markdown("### Thống kê sự cố trong năm qua (Sheet Tong hop):")
            for doc in summary_docs:
                st.write(doc)
        else:
            st.write("Không tìm thấy dữ liệu thống kê.")
    except Exception as e:
        st.write(f"Lỗi hiển thị thống kê: {str(e)}")

async def main():
    # Hàm chính chạy ứng dụng
    st.set_page_config(page_title="RAG Điện Viễn Thông (Nội bộ)", layout="wide")
    setup_session_state()
    if not handle_authentication():
        return

    st.title("📄 Hệ thống truy vấn tài liệu thông minh")
    has_db = check_existing_data()

    if has_db and "chroma" in st.session_state:
        doc_count = st.session_state.chroma.collection.count()
        logger.info(f"Số tài liệu hiện có: {doc_count}")

    st.sidebar.header("Cấu hình")
    
    # Khởi tạo giá trị mặc định cho llm_type và role nếu chưa có
    if "llm_type" not in st.session_state:
        st.session_state.llm_type = "ollama"
    if "role" not in st.session_state:
        st.session_state.role = "Expert"
    if "similarity_threshold" not in st.session_state:
        st.session_state.similarity_threshold = SIMILARITY_THRESHOLD_DEFAULT

    # Sử dụng st.form để nhóm các lựa chọn và tránh rerender không cần thiết
    with st.sidebar.form(key="config_form"):
        llm_type = st.radio(
            "Chọn mô hình LLM:",
            ["ollama", "openai"],
            format_func=lambda x: "Ollama Llama3.2 (Offline)" if x == "ollama" else "OpenAI GPT-4o-mini (Online)",
            index=["ollama", "openai"].index(st.session_state.llm_type)
        )
        role = st.radio(
            "Mức độ chi tiết:",
            ["Beginner", "Expert", "PhD"],
            index=["Beginner", "Expert", "PhD"].index(st.session_state.role)
        )
        similarity_threshold = st.slider("Ngưỡng tương đồng:", 0.0, 1.0, st.session_state.similarity_threshold, step=0.05)
        submit_config = st.form_submit_button("Áp dụng cấu hình")

    # Chỉ cập nhật session_state khi nhấn nút "Áp dụng cấu hình"
    if submit_config:
        st.session_state.llm_type = llm_type
        st.session_state.role = role
        st.session_state.similarity_threshold = similarity_threshold
        display_message("Đã cập nhật cấu hình!", "info")

    if has_db:
        display_message("CSDL đã có dữ liệu.", "info")
        st.session_state.has_db_notified = False
    else:
        if not st.session_state.has_db_notified:
            display_message("Chưa có dữ liệu. Vui lòng tải tài liệu.", "warning")
            st.session_state.has_db_notified = True

    st.sidebar.subheader("Quản lý tài liệu")
    uploaded_files = st.sidebar.file_uploader(
        "Tải lên tài liệu",
        type=SUPPORTED_FILE_TYPES,
        accept_multiple_files=True
    )
    if uploaded_files:
        with st.spinner("Đang xử lý tài liệu…"):
            await process_files(uploaded_files)

    if has_db:
        st.sidebar.subheader("Tài liệu đã tải")
        display_documents()
        if st.sidebar.button("Xóa tất cả tài liệu"):
            delete_all_documents()

    st.sidebar.subheader("Quản lý cuộc trò chuyện")
    if st.sidebar.button("Trò chuyện mới", help="Bắt đầu cuộc trò chuyện mới", key="new_conversation", type="primary"):
        create_new_conversation()
        logger.info("Tạo cuộc trò chuyện mới.")

    if st.session_state.confirm_delete:
        col1, col2 = st.sidebar.columns(2)
        with col1:
            if st.button("Đồng ý xóa"):
                if create_new_conversation(confirmed=True):
                    logger.info("Xác nhận xóa.")
        with col2:
            if st.button("Hủy"):
                st.session_state.confirm_delete = False
                display_message("Hủy tạo cuộc trò chuyện mới.", "info")
                logger.info("Hủy xóa.")

    display_conversation_list()

    with st.expander("📖 Gợi ý câu hỏi", expanded=False):
        st.markdown("""
        Để nhận câu trả lời chính xác, dùng thuật ngữ viễn thông như MPĐ, ATS, Interlock. Ví dụ:
        - Sự cố mất một lộ điện lưới lộ nổi ở N6, cần làm gì?
        - Lỗi ACB 4000A tủ LV1 cấp tới ATS1, cách xử lý?
        - Điều hòa đẩy cảnh báo HP, các bước khắc phục?
        - Sự cố 2 lộ điện, MPĐ lỗi ở N4, giải pháp là gì?
        - UPS hỏng module công suất, cách xử lý?
        - Trong năm vừa qua có bao nhiêu sự cố AC?
        """)

    query_input = st.text_input("Nhập câu hỏi (tiếng Việt):", value=st.session_state.query_input_value, key="query_input")
    st.session_state.query_input_value = query_input

    if st.button("Gửi câu hỏi"):
        if query_input:
            with st.spinner("Đang tạo câu trả lời…"):
                await handle_query(query_input, st.session_state.llm_type, st.session_state.role, st.session_state.similarity_threshold)

    with st.expander("📊 Xem thống kê sự cố", expanded=False):
        display_summary_data()

    with st.expander("📚 Xem trích dẫn nguồn", expanded=False):
        if st.session_state.last_citations:
            for idx, citation in enumerate(st.session_state.last_citations, 1):
                st.markdown(f"#### Trích dẫn {idx}:")
                st.markdown(f"- **Nguồn**: {citation['source']}")
                st.markdown(f"- **Tài liệu**: {citation['filename']}")
                st.markdown(f"- **Sheet**: {citation['sheet_name']}")
                st.markdown(f"- **Situation ID**: {citation['situation_id']}")
                st.markdown(f"- **Trang**: {citation['page_number']}")
                st.markdown(f"- **Tương đồng**: {citation['score']:.4f}")
                st.markdown(f"- **Nội dung**:")
                st.write(citation['content'])
                st.markdown("---")
        else:
            st.write(f"Không có trích dẫn đạt ngưỡng tương đồng.")

    with st.expander("📜 Lịch sử truy vấn", expanded=False):
        display_query_history()
              
if __name__ == "__main__":
    asyncio.run(main())